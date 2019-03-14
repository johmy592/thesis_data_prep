import sys
from openpyxl import load_workbook


def keep_lower_only(terms):
    return list(set([t.lower() for t in terms]))

def keep_up_to_trigram(terms):
    return [t for t in terms if len(t.split()) < 4]

def write_terms_only(terms_list, write_filename):
    wf = open(write_filename,'w+')
    for w in terms_list:
        wf.write(w)
        wf.write('\n')
    wf.close()

if __name__ == "__main__":
    if(len(sys.argv) != 3):
        print("Usage:")
        print("python3 make_term_vocab.py <terms_path> <save_path>")
        sys.exit()

    terms_path = sys.argv[1]
    save_path = sys.argv[2]    

    wb = load_workbook(terms_path)
    sheet = wb[wb.sheetnames[0]]
    startrow, startcol = sheet["A2"].row, sheet["A2"].column
    term_list = [sheet.cell(row=i, column=startcol).value for i in range(startrow,sheet.max_row)]

    print("Total term count: ", len(term_list))

    terms_lower = keep_lower_only(term_list)
    terms_lower.sort()
    terms_lower = keep_up_to_trigram(terms_lower)
    print("Terms left after lowercasing and pruning n>3-grams: ", len(terms_lower))

    write_terms_only(terms_lower, save_path)
